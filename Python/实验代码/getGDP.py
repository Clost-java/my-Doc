import json
import time
import requests
import openpyxl
import warnings
from pyecharts import options as opts
from pyecharts.charts import Bar, Map, Timeline
import pandas as pd



#获取当前时间戳
def getTime():
    return int(round(time.time()*1000))
# 获取数据
def getData():
    url = 'https://data.stats.gov.cn/easyquery.htm?cn=E0103'


    headers = {
       "user-agent":'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 Edg/107.0.1418.56'
    }
           
    params= {'m': 'QueryData', 'dbcode': 'fsnd', 'rowcode': 'reg', 'colcode': 'sj',
                   'wds': '[{"wdcode":"zb","valuecode":"A020101"}]', 'dfwds': '[{"wdcode":"sj","valuecode": "LAST20"}]',
                   'k1': str(getTime())
                   }
    return json.loads(requests.get(url, headers=headers, verify=False, params=params).text)
#数据写入excel表
def formatData():
    rawData = getData()["returndata"]

    dataJson = rawData["datanodes"]
    nameJson = rawData["wdnodes"][1]["nodes"]
    yearJson = rawData["wdnodes"][2]["nodes"]
    dataList = []
    nameList = []
    yearList = []

    for item in dataJson:
        dataList.append(item["data"]["strdata"])
    for item in nameJson:
        nameList.append(item["cname"])
    for item in yearJson:
        yearList.append(item["cname"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions.width = 60

#第一行数据
    ws.cell(1,1,'地区')
    for i in range(2,len(yearList)):
        ws.cell(1,i,yearList[i-2])
#第一列数据
    for i in range(2,len(nameList)):
        ws.cell(i,1,nameList[i-2])
# 具体数据
    for i in range(2,len(nameList)):
        for j in range(2,len(yearList)):
            ws.cell(i,j,dataList[(len(yearList))*(i-2) + j - 2])

    wb.save('outPutFile/GdpData.xlsx')

#可视化数据
datas = pd.read_excel('outPutFile/GdpData.xlsx')
def get_gdp_bar1(datas):
    c = (
        Bar()
            .add_xaxis(datas['地区'].values.tolist())
            .add_yaxis('全国各省GDP(亿元)', datas['2021年'].values.tolist())
            .set_global_opts(
            title_opts=opts.TitleOpts(title='2021年全国各省GDP(亿元)'),
            datazoom_opts=[opts.DataZoomOpts(), opts.DataZoomOpts(type_='inside')],
        )
        .render('outPutFile/2021年全国各省GDP_Bar1.html')
    )

def get_gdp_bar2(datas):
    sort_info = datas.sort_values(by='2021年', ascending=True)
    c = (
        Bar()
            .add_xaxis(sort_info['地区'].values.tolist())
            .add_yaxis('全国各省GDP(亿元)', sort_info['2021年'].values.tolist())
            .reversal_axis()
            .set_series_opts(label_opts=opts.LabelOpts(position="right"))
            .set_global_opts(
            title_opts=opts.TitleOpts(title='2021年全国各省GDP(亿元)'),
            datazoom_opts=[opts.DataZoomOpts(is_show=True
            , orient='vertical')],

        )
        .render('outPutFile/2021年全国各省GDP_Bar2.html')
    )

def get_gdp_map1(datas):
    datas['地区'].replace(regex=True, inplace=True, to_replace=['省', '市', '维吾尔自治区', '回族自治区', '壮族自治区', '自治区'], value=r'')
    map = (
        Map()
        .add('全国各省GDP(亿元)', datas[['地区', '2021年']].values.tolist(), 'china')
        .set_global_opts(
            title_opts=opts.TitleOpts(title='2021年全国各省GDP(亿元)'),
            visualmap_opts=opts.VisualMapOpts(max_=110000),
        )
    )
    map.render('outPutFile/2021年全国各省GDP_Map.html')

def get_gdp_map2(datas):
    datas['地区'].replace(regex=True, inplace=True, to_replace=['省', '市', '维吾尔自治区', '回族自治区', '壮族自治区', '自治区'], value=r'')
    tl = Timeline()
    # 播放的速度，单位毫秒（ms）
    tl.add_schema(play_interval=300, symbol='emptydiamond')
    for i in range(2004, 2022):
        map0 = (
            Map()
            .add('全国各省GDP(亿元)', datas[['地区', str(i) + '年']].values.tolist(), 'china')
            .set_global_opts(
                title_opts=opts.TitleOpts(title='{}年全国各省GDP(亿元)'.format(i)),
                visualmap_opts=opts.VisualMapOpts(max_=110000,is_piecewise=True),
            )
        )
        tl.add(map0, '{}年'.format(i))
    tl.render('outPutFile/2004-2021年全国各省GDP.html')



warnings.filterwarnings('ignore')
formatData()
get_gdp_bar1(datas)
get_gdp_bar2(datas)
get_gdp_map1(datas)
get_gdp_map2(datas)



